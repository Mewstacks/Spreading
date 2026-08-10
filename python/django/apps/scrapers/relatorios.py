"""Sincronização automática de relatórios de comissão.

O usuário não envia CSV. Cada marketplace expõe um adapter que busca/normaliza
linhas de receita a partir da conta conectada. Os adapters foram isolados para que
os seletores/URLs dos portais possam evoluir sem mexer no dashboard ou ranking.
"""
from __future__ import annotations

import hashlib
import csv
import io
import math
import re
import unicodedata
from datetime import datetime
from dataclasses import dataclass
from datetime import timedelta

from django.conf import settings
from django.contrib.auth import get_user_model
from django.db import transaction
from django.db.models import Q
from django.utils import timezone

from apps.scrapers.models import ReceitaAfiliado, RelatorioSync
from apps.scrapers.eventos import log_event


class ReportSyncActionRequired(Exception):
    """A conta precisa ser conectada/reconectada pelo usuário."""


class ReportSyncError(Exception):
    """Falha operacional do sync."""


class ReportSyncNaoConfigurado(Exception):
    """A leitura automática deste portal não está disponível.

    Diferente de ReportSyncActionRequired: aqui não há ação do usuário que resolva —
    falta configuração/implementação nossa. Tratar os dois como a mesma coisa mandava
    o usuário "reconectar" uma conta que já estava conectada, para sempre.
    """


class ReportCellInvalid(ValueError):
    """Célula presente mas ilegível; nunca deve ser convertida para zero."""


class ReportCellEmpty(ValueError):
    """Célula esperada veio vazia; é diferente do número zero."""


class ReportPeriodMismatch(ReportSyncError):
    """O portal devolveu uma linha fora da janela cuja aplicação foi solicitada."""


class ParsedRows(list):
    def __init__(self, rows=(), *, seen=0, rejected=0, schema_fingerprint=""):
        super().__init__(rows)
        self.seen = seen
        self.rejected = rejected
        self.schema_fingerprint = schema_fingerprint


@dataclass
class ReportRow:
    marketplace: str
    data: object
    etiqueta: str = ""
    produto_nome: str = ""
    cliques: int = 0
    conversoes: int = 0
    pedidos: int = 0
    receita: float = 0.0
    comissao: float = 0.0
    periodo_inicio: object | None = None
    periodo_fim: object | None = None
    granularidade: str = "dia"


_SO_NUMERO = re.compile(r"[^\d,.\-]")
_MILHAR_PT = re.compile(r"^-?\d{1,3}(\.\d{3})+$")


def _num(value) -> float:
    """Converte uma célula de portal em float. 0.0 quando não há número.

    Os portais são pt-BR e devolvem texto formatado ('R$ 1.234,56', '12,50', '3,2%').
    float() direto engolia tudo isso como 0.0 — e como o sync gravava status "ok" do
    mesmo jeito, o dashboard exibia R$ 0,00 com selo verde de "sincronizado".
    """
    if isinstance(value, bool):
        raise ReportCellInvalid("booleano não é métrica")
    if isinstance(value, (int, float)):
        v = float(value)
        if not math.isfinite(v):
            raise ReportCellInvalid("número não finito")
        return v
    raw = str(value or "").replace("\xa0", " ").strip()
    texto = _SO_NUMERO.sub("", raw).strip()
    if not raw or raw in {"-", "—", "–", "."}:
        return 0.0
    if not texto:
        raise ReportCellInvalid("número ilegível")
    if "," in texto:
        # pt-BR: '.' é milhar, ',' é decimal.
        texto = texto.replace(".", "").replace(",", ".")
    elif _MILHAR_PT.match(texto):
        # '1.234' sem vírgula: milhar pt-BR, não decimal ('1.234' = mil duzentos e
        # trinta e quatro cliques). '1.5' cai fora daqui e segue sendo decimal.
        texto = texto.replace(".", "")
    try:
        v = float(texto)
    except (TypeError, ValueError):
        raise ReportCellInvalid("número ilegível")
    if not math.isfinite(v):
        raise ReportCellInvalid("número não finito")
    return v


def _num_typed(value) -> tuple[str, float | None]:
    if value is None or (isinstance(value, str) and not value.strip()):
        return "empty", None
    if isinstance(value, str) and value.strip() in {"-", "—", "–", "."}:
        return "empty", None
    try:
        return "valid", _num(value)
    except ReportCellInvalid:
        return "invalid", None


def _counter(value, field):
    number = _num(value)
    if number < 0 or not float(number).is_integer():
        raise ReportCellInvalid(f"{field}: contador negativo ou fracionário")
    return int(number)


def _digest(usuario, row: ReportRow) -> str:
    raw = "|".join([
        str(usuario.id), row.marketplace, str(row.data), row.etiqueta,
        row.produto_nome, row.granularidade,
        str(row.periodo_inicio or ""), str(row.periodo_fim or ""),
    ])
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def resumo_financeiro(usuario) -> dict:
    """Soma série diária; usa snapshots só enquanto uma base pré-migração existir."""
    from django.db.models import Min, Max, Q, Sum

    serie = ReceitaAfiliado.objects.filter(
        usuario=usuario, origem="auto", granularidade="dia")
    if serie.exists():
        return serie.aggregate(
            pedidos=Sum("pedidos"), receita=Sum("receita"), comissao=Sum("comissao"),
            cliques_mkt=Sum("cliques"), conversoes=Sum("conversoes"),
            periodo_inicio=Min("periodo_inicio"), periodo_fim=Max("periodo_fim"),
        )
    # Compatibilidade transitória para bases ainda não migradas. A migração marca
    # esses registros como legacy, portanto produção deixa de entrar aqui após deploy.
    ultimos = (ReceitaAfiliado.objects.filter(usuario=usuario, origem="auto")
               .values("marketplace").annotate(ultima=Max("data")))
    filtro = Q(pk__in=[])
    for linha in ultimos:
        filtro |= Q(marketplace=linha["marketplace"], data=linha["ultima"])
    return ReceitaAfiliado.objects.filter(usuario=usuario).filter(filtro).aggregate(
        pedidos=Sum("pedidos"), receita=Sum("receita"), comissao=Sum("comissao"),
        cliques_mkt=Sum("cliques"), conversoes=Sum("conversoes"),
        periodo_inicio=Min("periodo_inicio"), periodo_fim=Max("periodo_fim"),
    )


def _upsert_rows(usuario, rows: list[ReportRow]) -> tuple[int, int]:
    criadas = atualizadas = 0
    with transaction.atomic():
        for row in rows:
            metrics = (row.cliques, row.conversoes, row.pedidos, row.receita, row.comissao)
            if any(value is None for value in metrics):
                raise ReportSyncError(
                    f"{row.marketplace}: linha contém métrica vazia; nada foi persistido."
                )
            defaults = {
                "usuario": usuario,
                "marketplace": row.marketplace,
                "data": row.data,
                "etiqueta": row.etiqueta[:120],
                "produto_nome": row.produto_nome[:255],
                "cliques": _counter(row.cliques, "cliques"),
                "conversoes": _counter(row.conversoes, "conversoes"),
                "pedidos": _counter(row.pedidos, "pedidos"),
                "receita": _num(row.receita),
                "comissao": _num(row.comissao),
                "periodo_inicio": row.periodo_inicio,
                "periodo_fim": row.periodo_fim,
                "origem": "auto",
                "granularidade": "dia",
            }
            _, created = ReceitaAfiliado.objects.update_or_create(
                hash_origem=_digest(usuario, row),
                defaults=defaults,
            )
            criadas += int(created)
            atualizadas += int(not created)
    return criadas, atualizadas


class BaseReportAdapter:
    marketplace = ""

    def fetch(self, usuario, desde, ate) -> list[ReportRow]:
        raise NotImplementedError


class MercadoLivreReportAdapter(BaseReportAdapter):
    marketplace = "mercadolivre"

    def fetch(self, usuario, desde, ate) -> list[ReportRow]:
        preflight = report_prerequisites(usuario, self.marketplace)
        if not preflight["ok"]:
            exc = ReportSyncActionRequired if preflight["code"] == "session_missing" else ReportSyncNaoConfigurado
            raise exc(preflight["instruction"])
        url = preflight["url"]
        return _fetch_browser_report(usuario, self.marketplace, url, desde, ate)


class AmazonReportAdapter(BaseReportAdapter):
    marketplace = "amazon"

    def fetch(self, usuario, desde, ate) -> list[ReportRow]:
        preflight = report_prerequisites(usuario, self.marketplace)
        if not preflight["ok"]:
            exc = ReportSyncActionRequired if preflight["code"] == "session_missing" else ReportSyncNaoConfigurado
            raise exc(preflight["instruction"])
        return _fetch_browser_report(
            usuario, self.marketplace, preflight["url"], desde, ate,
        )


def report_prerequisites(usuario, marketplace):
    """Diagnóstico sem browser, próprio para UI e scheduler."""
    from apps.accounts.feature_flags import enabled_for_user
    from apps.scrapers.report_sessions import has_report_session

    marketplace = str(marketplace or "").lower()
    if marketplace == "mercadolivre":
        if not enabled_for_user("ML_BROWSER_REPORTS_ENABLED", usuario):
            return {"ok": False, "code": "feature_disabled", "url": "",
                    "instruction": "Habilite ML_BROWSER_REPORTS_ENABLED para esta organização."}
        url_setting = "ML_AFFILIATE_REPORT_URL"
        portal = "Mercado Livre"
    elif marketplace == "amazon":
        if not enabled_for_user("AMAZON_BROWSER_REPORTS_ENABLED", usuario):
            return {"ok": False, "code": "feature_disabled", "url": "",
                    "instruction": "Habilite AMAZON_BROWSER_REPORTS_ENABLED para esta organização."}
        url_setting = "AMAZON_ASSOCIATES_REPORT_URL"
        portal = "Amazon Associados"
    else:
        return {"ok": False, "code": "unsupported_marketplace", "url": "",
                "instruction": "Marketplace sem adaptador de relatórios."}
    url = str(getattr(settings, url_setting, "") or "").strip()
    if not url:
        return {"ok": False, "code": "url_missing", "url": "",
                "instruction": f"Defina {url_setting} com a página real do relatório."}
    if not has_report_session(usuario, marketplace):
        return {"ok": False, "code": "session_missing", "url": "",
                "instruction": f"Conecte a sessão exclusiva de relatórios do {portal}."}
    return {"ok": True, "code": "ready", "url": url, "instruction": ""}


def _login_detected(page):
    try:
        url = str(page.url or "").lower()
    except Exception:
        url = ""
    if any(marker in url for marker in ("/login", "/signin", "/ap/signin", "/lgz/")):
        return True
    return bool(page.locator(
        "input[type='password'], input[name*='password' i]"
    ).count())


def _apply_period(page, desde, ate):
    """Aplica o período e falha se o portal não oferecer controles comprováveis."""
    start_selectors = (
        "input[name*='start' i]", "input[name*='inicio' i]",
        "input[aria-label*='início' i]", "input[placeholder*='início' i]",
    )
    end_selectors = (
        "input[name*='end' i]", "input[name*='fim' i]",
        "input[aria-label*='fim' i]", "input[placeholder*='fim' i]",
    )

    def first(selectors):
        for selector in selectors:
            locator = page.locator(selector)
            if locator.count():
                return locator.first
        return None

    start, end = first(start_selectors), first(end_selectors)
    if start is None or end is None:
        raise ReportSyncError(
            "Controles de período não reconhecidos; o relatório não foi sincronizado."
        )
    start.fill(desde.strftime("%Y-%m-%d"))
    end.fill(ate.strftime("%Y-%m-%d"))
    applied = False
    for selector in (
        "button:has-text('Aplicar')", "button:has-text('Filtrar')",
        "button:has-text('Atualizar')", "button[type='submit']",
    ):
        control = page.locator(selector)
        if control.count():
            control.first.click(timeout=2000)
            applied = True
            break
    if not applied:
        raise ReportSyncError(
            "Ação para aplicar o período não foi reconhecida; nada foi persistido."
        )
    try:
        page.wait_for_load_state("networkidle", timeout=10000)
    except Exception:
        pass


def _fetch_browser_report(usuario, marketplace: str, url: str, desde, ate) -> list[ReportRow]:
    """Executor browser-first.

    Este primeiro contrato espera que o portal exponha uma tabela HTML de relatório.
    Ele é propositalmente conservador: se a tabela ou sessão não estiver clara, falha
    com ação explícita em vez de inventar receita.
    """
    from playwright.sync_api import sync_playwright

    # Tanto ML quanto Amazon leem o relatório da SESSÃO DE RELATÓRIO cifrada
    # (report_sessions), não da sessão do site principal. Para o ML isso separa a
    # sessão do portal de afiliados (comissão) da sessão do Link Builder.
    from apps.scrapers.report_sessions import load_report_state, registrar_veredito
    from apps.accounts.models import organization_for_user
    from apps.scrapers.carga import operacao_pesada
    storage_state = load_report_state(usuario, marketplace)
    organization = organization_for_user(usuario)
    resource = (
        f"amazon_report_session:{organization.pk}"
        if marketplace == "amazon" else f"ml_report_session:{organization.pk}"
    )

    try:
        if not storage_state:
            raise ReportSyncActionRequired(f"Conecte o portal de relatórios {marketplace}.")
        with operacao_pesada(
            resource_key="django_chromium", owner_kind="reports",
            organization=organization,
        ) as chromium_acquired:
            if not chromium_acquired:
                raise ReportSyncError("Chromium ocupado por outra tarefa; tente novamente.")
            with operacao_pesada(
                resource_key=resource, owner_kind="reports",
                organization=organization,
            ) as session_acquired:
                if not session_acquired:
                    raise ReportSyncError("Sessão de relatório ocupada por outra tarefa.")
                with sync_playwright() as p:
                    browser = p.chromium.launch(headless=True)
                    try:
                        context = browser.new_context(storage_state=storage_state)
                        page = context.new_page()
                        page.goto(url, wait_until="domcontentloaded", timeout=45000)
                        if _login_detected(page):
                            raise ReportSyncActionRequired(
                                f"Sessão de relatórios {marketplace} expirada. Reconecte a conta."
                            )
                        _apply_period(page, desde, ate)
                        exported = _download_delimited_report(page)
                        rows = (_parse_export(
                            exported[1], exported[0], marketplace, desde, ate,
                        ) if exported is not None else _extract_paginated_table_rows(
                            page, marketplace, desde, ate))
                    finally:
                        browser.close()
            # O sync é o ÚNICO fluxo que usa a sessão de relatórios de verdade, então
            # é o único que sabe se ela vale. A tela lê este veredito — antes ela só
            # checava se o arquivo existia, e por isso mostrava verde para sempre.
            registrar_veredito(usuario, marketplace, "conectado")
            return rows
    except ReportSyncActionRequired as exc:
        # Suspeita, não sentença: registra e deixa a acumulação decidir. Uma falha
        # isolada (portal instável, layout mudado) não pode desconectar a conta.
        registrar_veredito(usuario, marketplace, "suspeito", str(exc))
        raise
    except Exception as exc:
        # A exceção de Playwright/requests pode carregar URL com query, headers ou
        # fragmentos do portal. O encadeamento preserva a causa para tooling local,
        # mas banco, UI e log recebem somente a categoria segura.
        raise ReportSyncError(
            f"{marketplace}: falha operacional ao ler o relatório automático "
            f"({type(exc).__name__})."
        ) from exc


def _header_key(texto: str) -> str:
    normal = unicodedata.normalize("NFKD", str(texto or "")).encode(
        "ascii", "ignore").decode("ascii").lower()
    return re.sub(r"[^a-z0-9]", "", normal)


_HEADERS = {
    "data": {"data", "date", "dia"},
    "etiqueta": {"etiqueta", "tag", "trackingid", "idderastreamento"},
    "produto": {"produto", "item", "nomeproduto"},
    "cliques": {"cliques", "clicks"},
    "pedidos": {"pedidos", "itenspedidos", "orders"},
    "conversoes": {"conversoes", "conversions", "itensconvertidos"},
    "receita": {"receita", "vendas", "faturamento", "revenue"},
    "comissao": {"comissao", "ganhos", "earnings", "commission"},
}
_REQUIRED_METRICS = frozenset({
    "cliques", "conversoes", "pedidos", "receita", "comissao",
})


def _header_indices(headers) -> dict[str, int]:
    indices = {}
    for campo, aliases in _HEADERS.items():
        for idx, header in enumerate(headers):
            if _header_key(header) in aliases:
                indices[campo] = idx
                break
    return indices


def _require_metric_schema(indices, marketplace, formato):
    missing = sorted(_REQUIRED_METRICS - set(indices))
    if missing:
        raise ReportSyncError(
            f"{marketplace}: {formato} sem coluna(s) obrigatória(s): "
            f"{', '.join(missing)}. Nenhuma métrica foi persistida."
        )


def _rows_from_cells(cells, indices, marketplace: str, desde, ate) -> ReportRow:
    def get(campo, default=""):
        pos = indices.get(campo)
        return cells[pos] if pos is not None and pos < len(cells) else default
    parsed = {}
    for campo in ("cliques", "conversoes", "pedidos", "receita", "comissao"):
        state, value = _num_typed(get(campo))
        if state == "empty":
            raise ReportCellEmpty(f"{campo}: célula vazia")
        if state == "invalid":
            raise ReportCellInvalid(f"{campo}: célula ilegível")
        parsed[campo] = value
    data = _date(get("data"), None) if "data" in indices else ate
    if data is None:
        raise ReportCellInvalid("data ilegível")
    if data < desde or data > ate:
        raise ReportPeriodMismatch(
            f"{marketplace}: o portal devolveu data fora do período solicitado."
        )
    return ReportRow(
        marketplace=marketplace, data=data, etiqueta=get("etiqueta"),
        produto_nome=get("produto"),
        cliques=_counter(parsed["cliques"], "cliques"),
        conversoes=_counter(parsed["conversoes"], "conversoes"),
        pedidos=_counter(parsed["pedidos"], "pedidos"),
        receita=parsed["receita"], comissao=parsed["comissao"],
        periodo_inicio=desde, periodo_fim=ate,
        granularidade="dia",
    )


def _parse_delimited_report(content: bytes, marketplace: str, desde, ate) -> list[ReportRow]:
    """Lê CSV/TSV por cabeçalho, aceitando exportações pt-BR e UTF-8 com BOM."""
    text = content.decode("utf-8-sig", errors="replace")
    try:
        dialect = csv.Sniffer().sniff(text[:4096], delimiters=",;\t")
    except csv.Error:
        dialect = csv.excel_tab if "\t" in text.partition("\n")[0] else csv.excel
    reader = csv.reader(io.StringIO(text), dialect)
    try:
        headers = next(reader)
    except StopIteration:
        raise ReportSyncError(f"{marketplace}: exportação vazia.")
    indices = _header_indices(headers)
    if not indices:
        raise ReportSyncError(f"{marketplace}: cabeçalhos da exportação não reconhecidos.")
    _require_metric_schema(indices, marketplace, "exportação")
    rows, seen, rejected = [], 0, 0
    for cells in reader:
        if not any(str(cell).strip() for cell in cells):
            continue
        seen += 1
        try:
            rows.append(_rows_from_cells(cells, indices, marketplace, desde, ate))
        except (ReportCellInvalid, ReportCellEmpty):
            rejected += 1
    if seen and not rows:
        raise ReportSyncError(
            f"{marketplace}: {seen} linha(s) vistas e nenhuma métrica legível."
        )
    fingerprint = hashlib.sha256(
        "|".join(_header_key(h) for h in headers).encode("utf-8")
    ).hexdigest()
    return ParsedRows(rows, seen=seen, rejected=rejected,
                      schema_fingerprint=fingerprint)


def _parse_xlsx_report(content: bytes, marketplace: str, desde, ate) -> list[ReportRow]:
    """Lê XLSX em modo read-only; fórmulas usam somente o valor calculado salvo."""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ReportSyncNaoConfigurado(
            "Instale openpyxl para ler exportações XLSX de relatórios."
        ) from exc
    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise ReportSyncError(f"{marketplace}: arquivo XLSX ilegível.") from exc
    try:
        sheet = workbook.active
        iterator = sheet.iter_rows(values_only=True)
        try:
            headers = ["" if value is None else str(value) for value in next(iterator)]
        except StopIteration:
            raise ReportSyncError(f"{marketplace}: exportação XLSX vazia.")
        indices = _header_indices(headers)
        if not indices:
            raise ReportSyncError(f"{marketplace}: cabeçalhos XLSX não reconhecidos.")
        _require_metric_schema(indices, marketplace, "XLSX")
        rows, seen, rejected = [], 0, 0
        for raw in iterator:
            if not any(value not in (None, "") for value in raw):
                continue
            seen += 1
            try:
                rows.append(_rows_from_cells(list(raw), indices, marketplace, desde, ate))
            except (ReportCellInvalid, ReportCellEmpty):
                rejected += 1
        if seen and not rows:
            raise ReportSyncError(
                f"{marketplace}: {seen} linha(s) XLSX e nenhuma métrica legível."
            )
        fingerprint = hashlib.sha256(
            "|".join(_header_key(h) for h in headers).encode("utf-8")
        ).hexdigest()
        return ParsedRows(rows, seen=seen, rejected=rejected,
                          schema_fingerprint=fingerprint)
    finally:
        workbook.close()


def _parse_export(content: bytes, filename: str, marketplace: str, desde, ate):
    lower = str(filename or "").lower()
    if lower.endswith(".xlsx") or content[:4] == b"PK\x03\x04":
        return _parse_xlsx_report(content, marketplace, desde, ate)
    return _parse_delimited_report(content, marketplace, desde, ate)


def _download_delimited_report(page) -> tuple[str, bytes] | None:
    """Prefere uma exportação do portal sem assumir um seletor específico.

    Portais mudam ids e estruturas com frequência, mas a ação costuma preservar uma
    palavra de intenção. Se ela abrir menu, gerar XLSX ou não existir, o adapter cai
    de forma segura para o parser DOM — nunca para uma URL global de relatório.
    """
    controls = page.locator("a, button")
    for index in range(controls.count()):
        control = controls.nth(index)
        try:
            label = control.inner_text(timeout=300).strip()
        except Exception:
            continue
        if not re.search(r"(?:export|baixar|download).*(?:csv|tsv|xlsx|excel)|(?:csv|tsv|xlsx|excel).*(?:export|baixar|download)", label, re.I):
            continue
        try:
            with page.expect_download(timeout=2500) as event:
                control.click(timeout=1000)
            download = event.value
            name = (download.suggested_filename or "").lower()
            if not name.endswith((".csv", ".tsv", ".txt", ".xlsx")):
                continue
            path = download.path()
            if path:
                with open(path, "rb") as handle:
                    return name, handle.read()
        except Exception:
            continue
    return None


def _date(value, fallback):
    if hasattr(value, "date") and not isinstance(value, str):
        try:
            return value.date()
        except (TypeError, ValueError):
            pass
    texto = str(value or "").strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(texto, fmt).date()
        except ValueError:
            pass
    return fallback


def _extract_table_rows(page, marketplace: str, desde, ate) -> list[ReportRow]:
    if page.locator("input[type='password'], input[name*='password' i]").count():
        raise ReportSyncActionRequired(
            f"Sessão de relatórios {marketplace} expirada. Reconecte a conta."
        )
    header_locator = page.locator("table thead th")
    legacy_fixture = False
    try:
        headers = [header_locator.nth(i).inner_text().strip()
                   for i in range(header_locator.count())]
    except AttributeError:
        # Compatibilidade do adapter anterior e de dumps históricos sem thead. O
        # browser real só aceita o caminho por cabeçalhos logo abaixo.
        headers, legacy_fixture = [], True
    indices = _header_indices(headers)
    if not indices:
        if not legacy_fixture:
            raise ReportSyncError(f"{marketplace}: colunas de métricas não reconhecidas.")
        indices = {"etiqueta": 0, "produto": 1, "cliques": 2,
                   "conversoes": 3, "pedidos": 4, "receita": 5, "comissao": 6}
    table_rows = page.locator("table tbody tr")
    count = table_rows.count()
    if count == 0:
        raise ReportSyncError(
            f"{marketplace}: relatório sem tabela detectável; parser precisa ser ajustado."
        )
    _require_metric_schema(indices, marketplace, "tabela HTML")
    out: list[ReportRow] = []
    rejected = 0
    for idx in range(count):
        cells = [
            table_rows.nth(idx).locator("td").nth(i).inner_text(timeout=1000).strip()
            for i in range(table_rows.nth(idx).locator("td").count())
        ]
        if not cells:
            continue
        try:
            out.append(_rows_from_cells(cells, indices, marketplace, desde, ate))
        except (ReportCellInvalid, ReportCellEmpty):
            rejected += 1
    if count and not out:
        raise ReportSyncError(
            f"{marketplace}: {count} linha(s) lidas e nenhuma métrica legível."
        )
    fingerprint = hashlib.sha256(
        "|".join(_header_key(h) for h in headers).encode("utf-8")
    ).hexdigest()
    return ParsedRows(out, seen=count, rejected=rejected,
                      schema_fingerprint=fingerprint)


def _extract_paginated_table_rows(page, marketplace: str, desde, ate, max_pages=100):
    rows, seen, rejected, schemas = [], 0, 0, []
    page_fingerprints = set()
    for _page_number in range(max_pages):
        current = _extract_table_rows(page, marketplace, desde, ate)
        raw = "|".join(
            f"{row.data}:{row.etiqueta}:{row.produto_nome}:{row.cliques}:{row.pedidos}:"
            f"{row.receita}:{row.comissao}" for row in current
        )
        fingerprint = hashlib.sha256(raw.encode("utf-8")).hexdigest()
        if fingerprint in page_fingerprints:
            raise ReportSyncError(
                f"{marketplace}: paginação repetiu a página anterior; dados não persistidos."
            )
        page_fingerprints.add(fingerprint)
        rows.extend(current)
        seen += current.seen
        rejected += current.rejected
        schemas.append(current.schema_fingerprint)

        next_control = None
        for selector in (
            "a[rel='next']", "button[aria-label*='próxima' i]",
            "button[aria-label*='next' i]", "a:has-text('Próxima')",
        ):
            candidate = page.locator(selector)
            if candidate.count():
                next_control = candidate.first
                break
        if next_control is None:
            break
        disabled = (
            next_control.get_attribute("disabled") is not None
            or str(next_control.get_attribute("aria-disabled") or "").lower() == "true"
        )
        if disabled:
            break
        next_control.click(timeout=2000)
        try:
            page.wait_for_load_state("networkidle", timeout=10000)
        except Exception:
            pass
    else:
        raise ReportSyncError(
            f"{marketplace}: paginação excedeu o limite seguro de {max_pages} páginas."
        )
    schema = hashlib.sha256("|".join(schemas).encode("utf-8")).hexdigest()
    return ParsedRows(rows, seen=seen, rejected=rejected, schema_fingerprint=schema)


ADAPTERS = {
    "mercadolivre": MercadoLivreReportAdapter(),
    "amazon": AmazonReportAdapter(),
}


def sync_marketplace(usuario, marketplace: str, dias: int = 14) -> RelatorioSync:
    marketplace = (marketplace or "").lower()
    if marketplace not in ADAPTERS:
        raise ReportSyncError(f"Marketplace inválido: {marketplace}")
    agora = timezone.now()
    ate = timezone.localdate()
    desde = ate - timedelta(days=max(1, dias))
    sync, _ = RelatorioSync.objects.get_or_create(
        usuario=usuario, marketplace=marketplace)
    preflight = report_prerequisites(usuario, marketplace)
    log_event("relatorios", "sync_started", f"Iniciando sync {marketplace}.",
              usuario=usuario, contexto={"marketplace": marketplace, "dias": dias})
    sync.status = "rodando"
    sync.ultimo_inicio = agora
    sync.erro = ""
    sync.prerequisite_code = preflight["code"]
    sync.save(update_fields=[
        "status", "ultimo_inicio", "erro", "prerequisite_code", "atualizado_em",
    ])
    try:
        rows = ADAPTERS[marketplace].fetch(usuario, desde, ate)
        criadas, atualizadas = _upsert_rows(usuario, rows)
    except ReportSyncNaoConfigurado as exc:
        sync.status = "nao_configurado"
        sync.erro = str(exc)[:500]
        sync.ultimo_fim = timezone.now()
        sync.prerequisite_code = preflight["code"]
        # Sem retry curto: não é falha transitória, é uma feature que não existe.
        sync.proxima_execucao = timezone.now() + timedelta(days=1)
        sync.save()
        log_event("relatorios", "sync_nao_configurado", str(exc), level="info",
                  usuario=usuario, contexto={"marketplace": marketplace})
        return sync
    except ReportSyncActionRequired as exc:
        sync.status = "acao"
        sync.erro = str(exc)[:500]
        sync.ultimo_fim = timezone.now()
        sync.prerequisite_code = "session_expired" if preflight["ok"] else preflight["code"]
        sync.proxima_execucao = timezone.now() + timedelta(hours=6)
        sync.save()
        log_event("relatorios", "sync_action_required", str(exc), level="warning",
                  usuario=usuario, contexto={"marketplace": marketplace})
        return sync
    except Exception as exc:
        sync.status = "erro"
        safe_error = (
            f"{marketplace}: falha operacional durante a sincronização "
            f"({type(exc).__name__})."
        )
        sync.erro = safe_error
        sync.ultimo_fim = timezone.now()
        sync.prerequisite_code = "operational_failure"
        sync.proxima_execucao = timezone.now() + timedelta(hours=6)
        sync.save()
        log_event("relatorios", "sync_failed", safe_error, level="error",
                  usuario=usuario, contexto={"marketplace": marketplace}, exc=exc)
        return sync

    sync.status = "ok"
    sync.ultimo_fim = timezone.now()
    sync.ultimo_sucesso = sync.ultimo_fim
    sync.proxima_execucao = timezone.now() + timedelta(hours=6)
    sync.registros_criados = criadas
    sync.registros_atualizados = atualizadas
    sync.erro = ""
    sync.prerequisite_code = "ready"
    sync.schema_fingerprint = str(getattr(rows, "schema_fingerprint", "") or "")[:64]
    sync.linhas_vistas = int(getattr(rows, "seen", len(rows)))
    sync.linhas_aceitas = len(rows)
    sync.linhas_rejeitadas = int(getattr(rows, "rejected", 0))
    sync.periodo_aplicado_inicio = desde
    sync.periodo_aplicado_fim = ate
    sync.save()
    log_event(
        "relatorios", "sync_ok", f"{marketplace}: sync concluído.",
        usuario=usuario,
        contexto={"marketplace": marketplace, "criadas": criadas, "atualizadas": atualizadas},
    )
    return sync


def sync_user_reports(usuario, marketplace: str | None = None) -> list[RelatorioSync]:
    marketplaces = [marketplace] if marketplace else list(ADAPTERS)
    return [sync_marketplace(usuario, m) for m in marketplaces]


def sync_due_reports(limit: int = 20) -> list[RelatorioSync]:
    agora = timezone.now()
    # A fila é o próprio RelatorioSync, não o primeiro N de usuários. Assim quem
    # está vencido há mais tempo sempre avança, inclusive acima de vinte contas.
    for user in get_user_model().objects.filter(is_active=True, perfil__bloqueado=False):
        for marketplace in ADAPTERS:
            RelatorioSync.objects.get_or_create(usuario=user, marketplace=marketplace)
    pendentes = (RelatorioSync.objects.filter(Q(proxima_execucao__isnull=True) | Q(proxima_execucao__lte=agora))
                 .select_related("usuario", "usuario__perfil")
                 .filter(usuario__is_active=True, usuario__perfil__bloqueado=False)
                 .order_by("proxima_execucao", "pk")[:limit])
    resultados = []
    for sync in pendentes:
        resultados.append(sync_marketplace(sync.usuario, sync.marketplace))
    return resultados
